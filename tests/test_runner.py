"""
Тесты для price-comparison skill.

Запуск:
    python3 -m pytest tests/ -v
    или
    python3 tests/test_runner.py
"""

import sys
import json
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / 'scripts'))

from runner_v2 import create_main_sheet
from eval import eval_price_comparison, is_numeric_price, extract_price, is_clickable_url

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install: pip install openpyxl")
    sys.exit(1)


def test_brand_extraction():
    """Проверка авто-определения бренда из названия."""
    wb = openpyxl.Workbook()
    results = [{"num": 1, "name": "Cisco C9200-24P-E"}]
    create_main_sheet(wb, results)
    ws = wb.active
    header = ws.cell(row=1, column=3).value
    assert "Cisco" in header, f"Expected 'Cisco' in header, got: {header}"
    print("✅ test_brand_extraction пройден")


def test_price_formatting():
    """Проверка форматирования цен в Excel."""
    wb = openpyxl.Workbook()
    results = [{
        "num": 1, "name": "Test",
        "price1": 106714, "supplier1": "shop.ru", "url1": "https://shop.ru",
        "price2": None, "supplier2": "", "url2": "",
        "analog": "", "analog_price": None, "analog_supplier": "", "analog_url": "",
        "comment": ""
    }]
    create_main_sheet(wb, results)
    ws = wb.active
    cell = ws.cell(row=2, column=3)
    assert cell.value == 106714
    assert '₽' in cell.number_format
    print("✅ test_price_formatting пройден")


def test_hyperlinks():
    """Проверка создания гиперссылок для URL."""
    wb = openpyxl.Workbook()
    results = [{
        "num": 1, "name": "Test",
        "price1": 100000, "supplier1": "shop.ru", "url1": "https://shop.ru/product",
        "price2": None, "supplier2": "", "url2": "",
        "analog": "", "analog_price": None, "analog_supplier": "", "analog_url": "",
        "comment": ""
    }]
    create_main_sheet(wb, results)
    ws = wb.active
    cell = ws.cell(row=2, column=5)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == "https://shop.ru/product"
    print("✅ test_hyperlinks пройден")


def test_analogs_list():
    """Проверка создания списка аналогов для исследования."""
    wb = openpyxl.Workbook()
    results = [
        {"num": 1, "name": "Cisco C9200", "analog": "Huawei S5735", "price1": 100000, "analog_price": 80000},
        {"num": 2, "name": "Beward SV3218", "analog": "Dahua IPC", "price1": 44600, "analog_price": 26500},
        {"num": 3, "name": "No Analog", "analog": ""}
    ]
    analogs = create_main_sheet(wb, results)
    assert len(analogs) == 2
    assert analogs[0]['original'] == "Cisco C9200"
    assert analogs[0]['analog'] == "Huawei S5735"
    print("✅ test_analogs_list пройден")


def test_numeric_price_parsing():
    """Проверка парсинга цен из разных форматов."""
    assert is_numeric_price(106714) == True
    assert is_numeric_price("106 714 ₽") == True
    assert is_numeric_price("106,714.50") == True
    assert is_numeric_price("дорого") == False
    assert extract_price("106 714 ₽") == 106714.0
    print("✅ test_numeric_price_parsing пройден")


def test_url_validation():
    """Проверка валидации URL."""
    assert is_clickable_url("https://shop.ru/product") == True
    assert is_clickable_url("http://example.com") == True
    assert is_clickable_url("") == False
    assert is_clickable_url("not-a-url") == False
    print("✅ test_url_validation пройден")


def test_full_pipeline():
    """Интеграционный тест полного pipeline."""
    with tempfile.TemporaryDirectory() as tmpdir:
        # Создаём тестовые данные
        results = [
            {
                "num": 1, "name": "Cisco C9200-24P-E",
                "price1": 106714, "supplier1": "shop.nag.ru", "url1": "https://shop.nag.ru/cisco-c9200",
                "price2": 182112, "supplier2": "ediscom.ru", "url2": "https://ediscom.ru/cisco-c9200",
                "analog": "Huawei S5735-S24P4X", "analog_price": 99194, "analog_supplier": "network.msk.ru", "analog_url": "https://network.msk.ru/huawei-s5735",
                "comment": "Аналог дешевле на 26%"
            },
            {
                "num": 2, "name": "Beward SV3218RBZ",
                "price1": 44600, "supplier1": "beward.ru", "url1": "https://beward.ru/sv3218",
                "price2": None, "supplier2": "", "url2": "",
                "analog": "Dahua IPC-HFW3541EP-AS", "analog_price": 26500, "analog_supplier": "dahuastore.ru", "analog_url": "https://dahuastore.ru/ipc-hfw3541ep",
                "comment": "Аналог дешевле на 40%, без PoE"
            }
        ]
        
        results_file = Path(tmpdir) / "results.json"
        with open(results_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False)
        
        # Запускаем runner
        import subprocess
        result = subprocess.run(
            [sys.executable, 'scripts/runner_v2.py', 'dummy.xlsx', str(results_file), tmpdir],
            cwd=Path(__file__).parent.parent,
            capture_output=True, text=True
        )
        assert result.returncode == 0, f"runner failed: {result.stderr}"
        
        # Проверяем созданные файлы
        excel_files = list(Path(tmpdir).glob("*.xlsx"))
        assert len(excel_files) > 0, "Excel file not created"
        
        # Проверяем eval
        status, errors, warnings, stats = eval_price_comparison(excel_files[0])
        assert status != "FAIL", f"FAIL checks found: {errors[:3]}"
        print(f"✅ test_full_pipeline пройден (status: {status}, WARN: {len(warnings)})")


if __name__ == '__main__':
    test_brand_extraction()
    test_price_formatting()
    test_hyperlinks()
    test_analogs_list()
    test_numeric_price_parsing()
    test_url_validation()
    test_full_pipeline()
    print("\n🎉 Все тесты пройдены!")
