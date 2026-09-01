#!/usr/bin/env python3
"""
Matrix Builder v6.1 — Combines main results + sub-agent research into final Excel.

Usage:
    python3 matrix_builder.py main_results.xlsx analogs_dir/ alternatives_dir/ output.xlsx

Where analogs_dir/ contains JSON files from sub-agents:
    analog_matrix_[original]_[analog].json
    
And alternatives_dir/ contains:
    alternative_matrix_[original]_[alternative].json
"""

import sys
import os
import json
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_main_results(filepath):
    """Load the main results workbook."""
    wb = openpyxl.load_workbook(filepath)
    return wb


def load_matrix_json(filepath):
    """Load a single comparison matrix from JSON."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_matrix_sheet(wb, original_name, compare_name, matrix_data, sheet_type='analog'):
    """Create a comparison matrix sheet for analog or alternative."""
    original_brand = original_name.split()[0] if original_name else "Оригинал"
    compare_brand = compare_name.split()[0] if compare_name else "Сравнение"
    
    if sheet_type == 'alternative':
        sheet_name = f"Матрица Альтернатива {compare_brand}"
    else:
        sheet_name = f"Матрица Аналог {compare_brand}"
    
    # Ensure unique sheet name (Excel limit 31 chars)
    base_name = sheet_name[:27]
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name}_{counter}"
        counter += 1
    
    ws = wb.create_sheet(title=sheet_name)
    
    # Styles
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    match_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    mismatch_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    neutral_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value='ТЕХНИЧЕСКАЯ СРАВНИТЕЛЬНАЯ МАТРИЦА')
    title_cell.font = Font(bold=True, size=14, color='2E75B6')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Subtitle
    ws.merge_cells('A2:E2')
    subtitle = ws.cell(row=2, column=1, value=f'Оригинал: {original_name}')
    subtitle.font = Font(bold=True, size=11)
    subtitle.alignment = Alignment(horizontal='left')
    
    ws.merge_cells('A3:E3')
    subtitle2 = ws.cell(row=3, column=1, value=f'{"Альтернатива" if sheet_type == "alternative" else "Аналог"}:   {compare_name}')
    subtitle2.font = Font(bold=True, size=11, color='C65911')
    subtitle2.alignment = Alignment(horizontal='left')
    
    # Headers
    headers = ['Категория', 'Параметр', 'Оригинал', 'Аналог' if sheet_type == 'analog' else 'Альтернатива', 'Статус']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Group data by category
    current_row = 6
    categories = {}
    for item in matrix_data:
        cat = item.get('category', 'Прочее')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)
    
    # Write data grouped by category
    for category, items in categories.items():
        # Category header row
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            cell.border = border
            cell.font = Font(bold=True)
        
        cat_cell = ws.cell(row=current_row, column=1, value=category)
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Parameters in this category
        for item in items:
            row = [
                '',
                item.get('parameter', ''),
                item.get('original', ''),
                item.get('analog', item.get('alternative', '')),
                item.get('status', '')
            ]
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color coding for status column
                if col == 5:
                    status_str = str(value)
                    if '✅' in status_str or 'Совпадает' in status_str:
                        cell.fill = match_fill
                    elif '🔴' in status_str or '❌' in status_str or 'Критично' in status_str:
                        cell.fill = mismatch_fill
                    elif '🟡' in status_str or 'Частично' in status_str or 'Некритично' in status_str:
                        cell.fill = warning_fill
                    else:
                        cell.fill = neutral_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    ws.row_dimensions[5].height = 30
    
    ws.freeze_panes = 'A6'
    
    # Add summary section
    summary_row = current_row + 2
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    summary_title = ws.cell(row=summary_row, column=1, value='ВЕРДИКТ ПО ЗАМЕНЕ')
    summary_title.font = Font(bold=True, size=12, color='2E75B6')
    
    # Count statuses
    statuses = [item.get('status', '') for item in matrix_data]
    match_count = sum(1 for s in statuses if '✅' in s)
    warning_count = sum(1 for s in statuses if '🟡' in s or '⚠️' in s)
    mismatch_count = sum(1 for s in statuses if '🔴' in s or '❌' in s)
    
    summary_row += 1
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    verdict = ws.cell(row=summary_row, column=1, 
                     value=f'Совпадений: {match_count} | Некритичных отличий: {warning_count} | Критичных отличий: {mismatch_count}')
    verdict.font = Font(size=11)
    
    return ws


def add_summary_sheet(wb, all_items):
    """Add a summary sheet with all analogs and alternatives overview."""
    ws = wb.create_sheet(title='Сводка', index=0)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Оригинал', 'Сравнение', 'Тип', 'Цена оригинала', 'Цена сравнения', 'Экономия', 'Рекомендация']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row_idx, item in enumerate(all_items, 2):
        row_data = [
            item.get('original', ''),
            item.get('compare', ''),
            item.get('type', ''),
            item.get('original_price', ''),
            item.get('compare_price', ''),
            item.get('savings', ''),
            item.get('recommendation', '')
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 25
    
    ws.freeze_panes = 'A2'
    return ws


def main():
    if len(sys.argv) < 5:
        print("Usage: python3 matrix_builder.py main_results.xlsx analogs_dir/ alternatives_dir/ output.xlsx")
        print("   or: python3 matrix_builder.py main_results.xlsx analogs_dir/ - output.xlsx  (no alternatives)")
        sys.exit(1)
    
    main_file = sys.argv[1]
    analogs_dir = sys.argv[2]
    alternatives_dir = sys.argv[3]
    output_file = sys.argv[4]
    
    # Load main results
    wb = load_main_results(main_file)
    
    all_items_summary = []
    
    # Process analogs (п.2)
    if analogs_dir != '-':
        pattern = os.path.join(analogs_dir, 'analog_matrix_*.json')
        matrix_files = glob.glob(pattern)
        print(f"Found {len(matrix_files)} analog matrix files")
        
        for matrix_file in matrix_files:
            data = load_matrix_json(matrix_file)
            original = data.get('original', 'Unknown')
            analog = data.get('analog', 'Unknown')
            matrix = data.get('matrix', [])
            
            if matrix:
                create_matrix_sheet(wb, original, analog, matrix, sheet_type='analog')
                print(f"Added analog matrix: {original} vs {analog}")
            
            all_items_summary.append({
                'original': original,
                'compare': analog,
                'type': 'Аналог (другая марка)',
                'original_price': data.get('original_price', ''),
                'compare_price': data.get('analog_price', ''),
                'savings': data.get('savings', ''),
                'recommendation': data.get('recommendation', '')
            })
    
    # Process alternatives (п.3)
    if alternatives_dir != '-':
        pattern = os.path.join(alternatives_dir, 'alternative_matrix_*.json')
        matrix_files = glob.glob(pattern)
        print(f"Found {len(matrix_files)} alternative matrix files")
        
        for matrix_file in matrix_files:
            data = load_matrix_json(matrix_file)
            original = data.get('original', 'Unknown')
            alternative = data.get('alternative', 'Unknown')
            matrix = data.get('matrix', [])
            
            if matrix:
                create_matrix_sheet(wb, original, alternative, matrix, sheet_type='alternative')
                print(f"Added alternative matrix: {original} vs {alternative}")
            
            all_items_summary.append({
                'original': original,
                'compare': alternative,
                'type': 'Альтернатива (та же марка)',
                'original_price': data.get('original_price', ''),
                'compare_price': data.get('alternative_price', ''),
                'savings': data.get('savings', ''),
                'recommendation': data.get('recommendation', '')
            })
    
    # Add summary sheet if we have items
    if all_items_summary:
        add_summary_sheet(wb, all_items_summary)
    
    # Save final workbook
    wb.save(output_file)
    print(f"Final workbook saved: {output_file}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")


if __name__ == '__main__':
    main()
