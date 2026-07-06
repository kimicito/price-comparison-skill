#!/usr/bin/env python3
"""
Eval скрипт для price-comparison skill v3.0.
Проверяет результат перед выдачей пользователю.

Usage:
    python3 eval.py /path/to/result.xlsx [input.xlsx]

Exit codes:
    0 = PASS
    1 = FAIL (критические ошибки)
    2 = WARN (есть замечания, но можно показать)
"""

import sys
import re
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install: pip install openpyxl")
    sys.exit(1)


def is_clickable_url(url):
    """Проверяет, что URL кликабельный."""
    if not url:
        return False
    url = str(url).strip()
    return url.startswith("http://") or url.startswith("https://")


def get_domain(url):
    """Извлекает домен из URL."""
    try:
        parsed = urlparse(str(url).strip())
        domain = parsed.netloc
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain
    except:
        return ""


def is_numeric_price(value):
    """Проверяет, что значение — число (int/float), не текст."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    # Пробуем распарсить строку
    cleaned = str(value).strip().replace(" ", "").replace("₽", "").replace("$", "")
    # Умная обработка запятых: если запятая между цифрами — разделитель тысяч, удаляем
    # Если запятая перед 1-2 цифрами в конце — десятичный разделитель, заменяем на точку
    import re
    # Сначала обрабатываем десятичные запятые (пример: 106,5 → 106.5)
    cleaned = re.sub(r'(\d),(\d{1,2})(?!\d)', r'\1.\2', cleaned)
    # Оставшиеся запятые — разделители тысяч, удаляем
    cleaned = cleaned.replace(",", "")
    try:
        float(cleaned)
        return True
    except:
        return False


def extract_price(value):
    """Извлекает числовое значение цены."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(" ", "").replace("₽", "").replace("$", "")
    import re
    cleaned = re.sub(r'(\d),(\d{1,2})(?!\d)', r'\1.\2', cleaned)
    cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned)
    except:
        return None


def eval_price_comparison(filepath, input_filepath=None):
    """Основная функция проверки."""
    
    errors = []
    warnings = []
    stats = {
        "total": 0,
        "with_price_1": 0,
        "with_price_2": 0,
        "with_url_1": 0,
        "with_url_2": 0,
        "missing_any_price": 0,
        "with_analog": 0,
        "analog_not_found": 0,
        "missing_comment": 0,
    }
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Результат сравнения'] if 'Результат сравнения' in wb.sheetnames else wb.active
    
    # Определяем заголовки по содержимому первой строки
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx
    
    def find_col(*variants):
        for v in variants:
            for h, idx in headers.items():
                if v in h:
                    return idx
        return None
    
    # Ищем колонки (гибко — по подстроке)
    col_num = find_col("№ п/п", "№", "номер")
    col_sku = find_col("наименование", "материал", "оборудование", "тмц")
    col_price_1 = find_col("цена 1", "цена1", "поставщик 1", "цена поставщика 1")
    col_url_1 = find_col("url 1", "url1", "сайт 1", "ссылка 1")
    col_supplier_1 = find_col("поставщик 1", "поставщик1")
    col_price_2 = find_col("цена 2", "цена2", "поставщик 2", "цена поставщика 2")
    col_url_2 = find_col("url 2", "url2", "сайт 2", "ссылка 2")
    col_supplier_2 = find_col("поставщик 2", "поставщик2")
    col_analog = find_col("аналог", "другая марка")
    col_analog_price = find_col("цена аналога")
    col_analog_url = find_col("url аналога", "сайт аналога")
    col_analog_supplier = find_col("поставщик аналога")
    col_comment = find_col("комментарий")
    
    # Проверка наличия обязательных колонок
    if not col_sku:
        errors.append("FAIL: Не найдена колонка 'Наименование ТМЦ'")
    if not col_price_1:
        errors.append("FAIL: Не найдена колонка 'Цена 1'")
    if not col_analog:
        errors.append("FAIL: Не найдена колонка 'Аналог (другая марка)'")
    if not col_comment:
        errors.append("FAIL: Не найдена колонка 'Комментарий'")
    
    if errors and len(errors) >= 3:  # Если критичных колонок не хватает
        return "FAIL", errors, warnings, stats
    
    # Проверка количества позиций vs input
    if input_filepath and Path(input_filepath).exists():
        try:
            wb_input = openpyxl.load_workbook(input_filepath)
            ws_input = wb_input.active
            input_count = ws_input.max_row - 1  # минус заголовок
            if abs(input_count - (ws.max_row - 1)) > 0:
                errors.append(f"FAIL: Количество позиций изменилось: input={input_count}, output={ws.max_row-1}")
        except Exception as e:
            warnings.append(f"WARN: Не удалось сравнить с input файлом: {e}")
    
    # Проверка каждой строки
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        sku_cell = row[col_sku - 1] if col_sku else None
        sku = str(sku_cell.value).strip()[:50] if sku_cell and sku_cell.value else f"строка_{row_idx}"
        
        stats["total"] += 1
        
        # --- Извлекаем значения ---
        price_1 = row[col_price_1 - 1].value if col_price_1 else None
        price_2 = row[col_price_2 - 1].value if col_price_2 else None
        url_1 = row[col_url_1 - 1].value if col_url_1 else None
        url_2 = row[col_url_2 - 1].value if col_url_2 else None
        supplier_1 = row[col_supplier_1 - 1].value if col_supplier_1 else None
        supplier_2 = row[col_supplier_2 - 1].value if col_supplier_2 else None
        analog = row[col_analog - 1].value if col_analog else None
        analog_price = row[col_analog_price - 1].value if col_analog_price else None
        analog_url = row[col_analog_url - 1].value if col_analog_url else None
        comment = row[col_comment - 1].value if col_comment else None
        
        has_price_1 = price_1 is not None and str(price_1).strip() != ""
        has_price_2 = price_2 is not None and str(price_2).strip() != ""
        has_analog = analog is not None and str(analog).strip() not in ["", "-", "Аналог не найден", "Нет прямого аналога"]
        
        if has_price_1:
            stats["with_price_1"] += 1
        if has_price_2:
            stats["with_price_2"] += 1
        if has_analog:
            stats["with_analog"] += 1
        if analog and "не найден" in str(analog).lower():
            stats["analog_not_found"] += 1
        
        # ==================== FAIL CHECKS ====================
        
        # FAIL 1: Комментарий обязателен
        if not comment or str(comment).strip() in ["", "-"]:
            errors.append(f"FAIL [{sku}]: Пустой комментарий. Каждая позиция должна иметь заполненный комментарий.")
            stats["missing_comment"] += 1
        
        # FAIL 2: Хотя бы одна цена оригинала
        if not has_price_1 and not has_price_2:
            stats["missing_any_price"] += 1
            errors.append(f"FAIL [{sku}]: Нет ни одной цены оригинала. Минимум 1 цена обязательна.")
            continue
        
        # FAIL 3: Цены — числовой формат
        if has_price_1 and not is_numeric_price(price_1):
            errors.append(f"FAIL [{sku}]: Цена 1 не является числом ({price_1}). Должна быть числовая ячейка.")
        if has_price_2 and not is_numeric_price(price_2):
            errors.append(f"FAIL [{sku}]: Цена 2 не является числом ({price_2}). Должна быть числовая ячейка.")
        if has_analog and analog_price is not None and str(analog_price).strip() != "" and not is_numeric_price(analog_price):
            errors.append(f"FAIL [{sku}]: Цена аналога не является числом ({analog_price}).")
        
        # FAIL 4: URL кликабельны
        if has_price_1 and not is_clickable_url(url_1):
            errors.append(f"FAIL [{sku}]: URL поставщика 1 не кликабелен ({url_1})")
        elif is_clickable_url(url_1):
            stats["with_url_1"] += 1
        
        if has_price_2 and not is_clickable_url(url_2):
            errors.append(f"FAIL [{sku}]: URL поставщика 2 не кликабелен ({url_2})")
        elif is_clickable_url(url_2):
            stats["with_url_2"] += 1
        
        # FAIL 5: Поставщики не дублируются
        if has_price_1 and has_price_2 and supplier_1 and supplier_2:
            dom1 = get_domain(url_1) if url_1 else str(supplier_1).strip().lower()
            dom2 = get_domain(url_2) if url_2 else str(supplier_2).strip().lower()
            if dom1 and dom2 and dom1 == dom2:
                errors.append(f"FAIL [{sku}]: Поставщики 1 и 2 — один домен ({dom1}). Нужны разные источники.")
        
        # FAIL 6: Аналог — другая марка
        if has_analog and sku:
            original_brand = str(sku).split()[0].lower() if sku else ""
            analog_brand = str(analog).split()[0].lower() if analog else ""
            if original_brand and analog_brand and original_brand == analog_brand:
                errors.append(f"FAIL [{sku}]: Аналог той же марки ({analog_brand}). Должен быть другой производитель.")
        
        # ==================== WARN CHECKS ====================
        
        # WARN 1: Цена аналога > оригинала → нужно объяснение
        if has_analog and has_price_1:
            p1 = extract_price(price_1)
            pa = extract_price(analog_price)
            if p1 and pa and pa >= p1:
                if comment and ("дороже" not in str(comment).lower() and "лучше" not in str(comment).lower() and "премиум" not in str(comment).lower()):
                    warnings.append(f"WARN [{sku}]: Аналог дороже оригинала ({pa:,.0f} ≥ {p1:,.0f}), но в комментарии нет объяснения.")
        
        # WARN 2: Экономия рассчитана (в комментарии есть сумма или %)
        if has_analog and has_price_1:
            p1 = extract_price(price_1)
            pa = extract_price(analog_price)
            if p1 and pa and pa < p1:
                economy = p1 - pa
                economy_pct = (economy / p1) * 100
                if comment:
                    comment_lower = str(comment).lower()
                    has_economy = any(word in comment_lower for word in ["экономия", "дешевле", "разница", "выгода", "%", "руб", "₽", "р."])
                    if not has_economy:
                        warnings.append(f"WARN [{sku}]: Не указана экономия в комментарии. Расчёт: {economy:,.0f} ₽ ({economy_pct:.0f}%).")
        
        # WARN 3: Только 1 цена оригинала
        if has_price_1 and not has_price_2:
            warnings.append(f"WARN [{sku}]: Только 1 цена оригинала. Цель: 2 цены от разных поставщиков.")
        
        # WARN 4: Цены отличаются подозрительно
        if has_price_1 and has_price_2:
            p1 = extract_price(price_1)
            p2 = extract_price(price_2)
            if p1 and p2 and p1 > 0 and p2 > 0:
                ratio = max(p1, p2) / min(p1, p2)
                if ratio > 10:
                    errors.append(f"FAIL [{sku}]: Цены оригинала отличаются в {ratio:.1f}x ({p1:,.0f} vs {p2:,.0f}). Проверьте артикул.")
                elif ratio > 3:
                    warnings.append(f"WARN [{sku}]: Цены отличаются в {ratio:.1f}x. Рекомендуется проверка ({p1:,.0f} vs {p2:,.0f}).")
        
        # WARN 5: Абсурдные цены
        if has_price_1:
            p1 = extract_price(price_1)
            if p1 and p1 < 10:
                warnings.append(f"WARN [{sku}]: Подозрительно низкая цена ({p1:,.0f} ₽). Проверьте единицу измерения.")
            if p1 and p1 > 10_000_000:
                warnings.append(f"WARN [{sku}]: Подозрительно высокая цена ({p1:,.0f} ₽). Проверьте единицу измерения.")
        
        # WARN 6: Есть аналог, но нет URL аналога
        if has_analog and analog_price is not None and not is_clickable_url(analog_url):
            warnings.append(f"WARN [{sku}]: У аналога нет кликабельного URL ({analog_url}).")
    
    # ==================== ПРОВЕРКА МАТРИЦ ====================
    matrix_sheets = [s for s in wb.sheetnames if s.startswith("Матрица")]
    
    # WARN 7: Если есть аналоги, но нет матриц
    if stats["with_analog"] > 0 and len(matrix_sheets) == 0:
        warnings.append(f"WARN: Найдено {stats['with_analog']} аналогов, но нет вкладок с матрицами сравнения.")
    
    # WARN 8: Критичные отличия в матрице → проверить комментарий
    for sheet_name in matrix_sheets:
        ws_matrix = wb[sheet_name]
        has_critical = False
        for row in ws_matrix.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell and "🔴" in str(cell):
                    has_critical = True
                    break
            if has_critical:
                break
        if has_critical:
            # Проверим, что в основном листе есть объяснение риска
            found_risk_comment = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                comment_cell = row[col_comment - 1] if col_comment and col_comment <= len(row) else None
                if comment_cell:
                    c = str(comment_cell).lower()
                    if any(word in c for word in ["критич", "риск", "опасн", "вниман", "⚠️", "🔴"]):
                        found_risk_comment = True
                        break
            if not found_risk_comment:
                warnings.append(f"WARN [{sheet_name}]: В матрице есть критичные отличия (🔴), но в комментариях нет упоминания рисков.")
    
    return ("FAIL" if errors else "WARN" if warnings else "PASS"), errors, warnings, stats


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 eval.py <result.xlsx> [input.xlsx]")
        sys.exit(1)
    
    filepath = Path(sys.argv[1])
    input_filepath = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    
    if not filepath.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)
    
    status, errors, warnings, stats = eval_price_comparison(filepath, input_filepath)
    
    # Вывод
    print(f"\n{'='*70}")
    print(f"EVAL RESULT: {status}")
    print(f"{'='*70}")
    print(f"\n📊 Статистика:")
    print(f"  Всего позиций: {stats['total']}")
    print(f"  С ценой 1 (оригинал): {stats['with_price_1']} ({stats['with_price_1']/max(stats['total'],1)*100:.0f}%)")
    print(f"  С ценой 2 (оригинал): {stats['with_price_2']} ({stats['with_price_2']/max(stats['total'],1)*100:.0f}%)")
    print(f"  С аналогом: {stats['with_analog']} ({stats['with_analog']/max(stats['total'],1)*100:.0f}%)")
    print(f"  Аналог не найден: {stats['analog_not_found']}")
    print(f"  Без комментария: {stats['missing_comment']}")
    print(f"  Без цен: {stats['missing_any_price']}")
    
    if errors:
        print(f"\n❌ FAIL ({len(errors)}):")
        for e in errors[:15]:
            print(f"  • {e}")
        if len(errors) > 15:
            print(f"  ... и ещё {len(errors) - 15}")
    
    if warnings:
        print(f"\n⚠️  WARN ({len(warnings)}):")
        for w in warnings[:15]:
            print(f"  • {w}")
        if len(warnings) > 15:
            print(f"  ... и ещё {len(warnings) - 15}")
    
    print(f"\n{'='*70}")
    
    if status == "PASS":
        print("✅ Готово к выдаче пользователю")
        sys.exit(0)
    elif status == "WARN":
        print("⚠️  Есть замечания, но можно показать")
        sys.exit(2)
    else:
        print("❌ Требуется доработка — отправь feedback агенту")
        sys.exit(1)


if __name__ == "__main__":
    main()
