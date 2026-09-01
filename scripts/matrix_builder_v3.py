#!/usr/bin/env python3
"""
Matrix Builder v3 — Combines main results + analog matrices into final Excel.

All analogs of the same type are placed on ONE sheet with sections.

Usage:
    python3 matrix_builder_v3.py main_results.xlsx analogs_dir/ output.xlsx

Where analogs_dir/ contains JSON files:
    analog_matrix_[original]_[analog]_[type].json
    
Type: 'other' or 'same' (brand)
"""

import sys
import os
import json
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# === DESIGN CONSTANTS ===
SECTION_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
REC_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

WRAP_ALIGN = Alignment(vertical='top', wrap_text=True, horizontal='left')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)


def load_main_results(filepath):
    """Load the main results workbook."""
    return openpyxl.load_workbook(filepath)


def load_analog_matrix(filepath):
    """Load a single analog comparison matrix from JSON."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def write_section(ws, start_row, section_data, num_cols=5):
    """Write a single analog section to worksheet with sources."""
    row = start_row
    
    # Section title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    title_cell = ws.cell(row=row, column=1, value=section_data.get('title', 'Аналог'))
    title_cell.fill = SECTION_FILL
    title_cell.font = SECTION_FONT
    title_cell.alignment = CENTER_ALIGN
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1
    
    # Sources row (if present)
    sources = section_data.get('sources', [])
    if sources:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        sources_text = 'Источники: ' + ' | '.join(sources)
        cell = ws.cell(row=row, column=1, value=sources_text)
        cell.fill = GRAY_FILL
        cell.font = Font(name='Calibri', size=8, color='0563C1', underline='single')
        cell.alignment = WRAP_ALIGN
        for c in range(1, num_cols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        # Make URLs clickable
        for i, src in enumerate(sources):
            if src.startswith('http'):
                # We can't make parts of merged cell clickable individually,
                # so we add a separate row with clickable links
                pass
        row += 1
    
    # Headers
    headers = section_data.get('headers', [])
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1
    
    # Data rows
    for row_data in section_data.get('rows', []):
        is_rec = False
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            cell.font = Font(name='Calibri', size=9)
            
            if col_num == 1 and isinstance(value, str) and 'РЕКОМЕНДАЦИЯ' in value:
                is_rec = True
            
            # Color coding for deviation column (usually col 4)
            if col_num == 4 and value and not is_rec:
                if '✅' in str(value):
                    cell.fill = GREEN_FILL
                    cell.font = Font(color='006100', bold=True)
                elif '🟡' in str(value):
                    cell.fill = YELLOW_FILL
                    cell.font = Font(color='9C5700', bold=True)
                elif '🔴' in str(value):
                    cell.fill = RED_FILL
                    cell.font = Font(color='9C0006', bold=True)
        
        # Merge recommendation row
        if is_rec:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            rec_cell = ws.cell(row=row, column=1)
            rec_cell.fill = REC_FILL
            rec_cell.font = Font(name='Calibri', size=10, bold=True)
            rec_cell.alignment = WRAP_ALIGN
        
        row += 1
    
    return row + 1  # +1 for empty row between sections


def create_analogs_sheet(wb, sheet_name, analogs_data):
    """Create a sheet with all analogs of one type (other or same brand)."""
    ws = wb.create_sheet(title=sheet_name)
    
    row = 1
    for section in analogs_data:
        cols = section.get('cols', 5)
        row = write_section(ws, row, section, num_cols=cols)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 30
    if hasattr(ws.column_dimensions, 'F'):
        ws.column_dimensions['F'].width = 24
    if hasattr(ws.column_dimensions, 'G'):
        ws.column_dimensions['G'].width = 30
    
    return ws


def build_matrix_json_for_analog(original, analog, analog_type='other'):
    """Build a matrix JSON structure for a single analog comparison."""
    # This would normally come from sub-agent research
    # Here we create a template structure
    return {
        'title': f'{original} → {analog}',
        'cols': 5,
        'headers': ['Параметр', 'Оригинал', 'Аналог', 'Отклонение', 'Влияние'],
        'rows': [
            ['Цена', '—', '—', '—', '—'],
            ['РЕКОМЕНДАЦИЯ', 'Требуется исследование субагентом', '', '', '']
        ]
    }


def main():
    if len(sys.argv) < 4:
        print("Usage: python3 matrix_builder_v3.py main_results.xlsx analogs_dir/ output.xlsx")
        print("\nanalogs_dir/ should contain JSON files:")
        print("  analog_matrix_[original]_[analog]_[type].json")
        sys.exit(1)
    
    main_file = sys.argv[1]
    analogs_dir = sys.argv[2]
    output_file = sys.argv[3]
    
    # Load main results
    wb = load_main_results(main_file)
    
    # Load all analog matrices
    other_brand_analogs = []
    same_brand_analogs = []
    
    if os.path.isdir(analogs_dir):
        for filepath in glob.glob(os.path.join(analogs_dir, 'analog_matrix_*.json')):
            data = load_analog_matrix(filepath)
            analog_type = data.get('type', 'other')
            if analog_type == 'same':
                same_brand_analogs.append(data)
            else:
                other_brand_analogs.append(data)
    
    # Create sheets
    if other_brand_analogs:
        create_analogs_sheet(wb, 'Аналоги другой марки', other_brand_analogs)
    
    if same_brand_analogs:
        create_analogs_sheet(wb, 'Аналоги той же марки', same_brand_analogs)
    
    # Save
    wb.save(output_file)
    print(f"✅ Final Excel saved: {output_file}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Other brand analogs: {len(other_brand_analogs)}")
    print(f"   Same brand analogs: {len(same_brand_analogs)}")


if __name__ == '__main__':
    main()
