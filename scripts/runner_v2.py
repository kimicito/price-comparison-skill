#!/usr/bin/env python3
"""
Price Comparison Runner v2 — Enhanced with Analog Research Support

Creates main results sheet and outputs analogs.json for sub-agent research.

Usage:
    python3 runner_v2.py input.xlsx results.json [output_dir]

Where results.json is a JSON file with price data from the agent:
[
  {
    "num": 1,
    "name": "Cisco C9200-24P-E",
    "price1": 106714,
    "supplier1": "shop.nag.ru",
    "url1": "...",
    "price2": 182112,
    "supplier2": "ediscom.ru",
    "url2": "...",
    "analog": "Huawei S5735-S24P4X",
    "analog_price": 99194,
    "analog_supplier": "network.msk.ru",
    "analog_url": "...",
    "comment": "..."
  }
]
"""

import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_main_sheet(wb, results):
    """Create the main results sheet with actual price data."""
    ws = wb.active
    ws.title = "Результат сравнения"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Авто-определение бренда из названия позиции (первое слово)
    brand = "Оригинал"
    if results and results[0].get('name'):
        # Извлекаем первое слово как бренд (Cisco, Beward, и т.д.)
        first_word = str(results[0]['name']).split()[0]
        if first_word and len(first_word) > 1:
            brand = first_word
    
    # Заголовки с динамическим брендом
    headers = [
        '№ п/п', 'Наименование ТМЦ по КЛ',
        f'Цена 1 (₽)\n{brand}, без НДС', 'Поставщик 1', 'URL 1',
        f'Цена 2 (₽)\n{brand}, без НДС', 'Поставщик 2', 'URL 2',
        'Аналог (другая марка)', 'Цена аналога (₽)', 'Поставщик аналога', 'URL аналога',
        'Комментарий'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    
    for row_idx, item in enumerate(results, 2):
        row_data = [
            item.get('num', ''), item.get('name', ''),
            item.get('price1', ''), item.get('supplier1', ''), item.get('url1', ''),
            item.get('price2', ''), item.get('supplier2', ''), item.get('url2', ''),
            item.get('analog', ''), item.get('analog_price', ''), item.get('analog_supplier', ''), item.get('analog_url', ''),
            item.get('comment', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col in [3, 6, 10] and isinstance(value, (int, float)):
                cell.number_format = '#,##0 "₽"'
            if col in [5, 8, 12] and value and str(value).startswith('http'):
                cell.hyperlink = value
                cell.font = Font(color='0563C1', underline='single')
    
    column_widths = [6, 35, 16, 18, 40, 16, 18, 40, 28, 16, 22, 40, 80]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.row_dimensions[1].height = 35
    for i in range(2, len(results) + 2):
        ws.row_dimensions[i].height = 100
    
    ws.freeze_panes = 'A2'
    
    analogs_to_research = []
    for item in results:
        if item.get('analog'):
            analogs_to_research.append({
                'original': item['name'],
                'analog': item['analog'],
                'original_price': item.get('price1'),
                'analog_price': item.get('analog_price'),
                'row_num': item.get('num')
            })
    
    return analogs_to_research


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 runner_v2.py input.xlsx results.json [output_dir]")
        print("\nresults.json format:")
        print(json.dumps([{
            "num": 1, "name": "Product Name",
            "price1": 100000, "supplier1": "supplier.ru", "url1": "https://...",
            "price2": null, "supplier2": "", "url2": "",
            "analog": "Analog Name", "analog_price": 80000, "analog_supplier": "...", "analog_url": "",
            "comment": ""
        }], indent=2, ensure_ascii=False))
        sys.exit(1)
    
    input_file = sys.argv[1]
    results_file = sys.argv[2]
    output_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.dirname(input_file) or '.'
    
    with open(results_file, 'r', encoding='utf-8') as f:
        results = json.load(f)
    
    print(f"Loaded {len(results)} items from {results_file}")
    
    wb = openpyxl.Workbook()
    analogs_to_research = create_main_sheet(wb, results)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    intermediate_file = os.path.join(output_dir, f'price_comparison_main_{timestamp}.xlsx')
    wb.save(intermediate_file)
    
    analogs_file = os.path.join(output_dir, f'analogs_to_research_{timestamp}.json')
    with open(analogs_file, 'w', encoding='utf-8') as f:
        json.dump(analogs_to_research, f, ensure_ascii=False, indent=2)
    
    print(f"Main results saved: {intermediate_file}")
    print(f"Analogs for research: {analogs_file}")
    print(f"Total analogs to research: {len(analogs_to_research)}")
    
    if analogs_to_research:
        print("\n--- ANALOGS_TO_RESEARCH ---")
        print(json.dumps(analogs_to_research, ensure_ascii=False))


if __name__ == '__main__':
    main()
