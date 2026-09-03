#!/usr/bin/env python3
"""
Price Comparison Runner v3 — Formula 2+1+1 + Design + Clickable Links

Creates main results sheet with:
  - 2 original prices (same brand, different suppliers)
  - 1 analog (different brand)
  - 1 alternative (same brand, different model)

All prices are clickable hyperlinks to product pages.

Usage:
    python3 runner_v3.py input.xlsx results.json [output_dir]

Where results.json format:
[
  {
    "num": 1,
    "name": "GIGALINK SFP GL-OT-SG07LC2-0850-0850-M",
    "price1": 925, "url1": "https://...", "supplier1": "gven.ru",
    "price2": 1570, "url2": "https://...", "supplier2": "chipdip.ru",
    "analog_brand": "SNR-SFP-SX",
    "analog_price": 700, "analog_url": "https://...", "analog_supplier": "shop.nag.ru",
    "alt_brand": "GL-OT-SG07LC2-I-M",
    "alt_price": 2391, "alt_url": "https://...", "alt_supplier": "svetelektro.net",
    "comment": "Согласовать SNR (другая марка)",
    "date": "2026-09-01"
  }
]
"""

import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from inline_eval import inline_eval_all


def is_clickable_url(url):
    """Проверяет, что URL кликабельный."""
    if not url:
        return False
    url = str(url).strip()
    return url.startswith("http://") or url.startswith("https://")


# === DESIGN CONSTANTS ===
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
PRICE_FONT = Font(name='Calibri', size=10, bold=True, color='0563C1', underline='single')
SHOP_FONT = Font(name='Calibri', size=8, color='666666', italic=True)
LINK_FONT = Font(name='Calibri', size=9, color='0563C1', underline='single')

REC_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
REC_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
REC_RED = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
WRAP_ALIGN = Alignment(vertical='top', wrap_text=True, horizontal='left')


def create_main_sheet(wb, results):
    """Create the main results sheet with 2+1+1 formula and graceful handling of missing data."""
    ws = wb.active
    ws.title = "Сводная таблица"
    
    # Headers
    headers = [
        '№', 'Наименование ТМЦ',
        'Цена 1 (₽)', 'Магазин 1', 'URL 1',
        'Цена 2 (₽)', 'Магазин 2', 'URL 2',
        'Аналог другой\nмарки (₽)', 'Магазин аналога', 'URL аналога',
        'Аналог той же\nмарки (₽)', 'Магазин альтерн.', 'URL альтерн.',
        'Рекомендация'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Data rows
    for r_idx, item in enumerate(results, 2):
        bg = GRAY_FILL if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        # №
        c = ws.cell(row=r_idx, column=1, value=item.get('num', ''))
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        
        # Name
        c = ws.cell(row=r_idx, column=2, value=item.get('name', ''))
        c.alignment = LEFT_ALIGN; c.border = THIN_BORDER; c.fill = bg
        c.font = Font(name='Calibri', size=9)
        
        # === Price 1 (REQUIRED) ===
        p1 = item.get('price1')
        url1 = item.get('url1', '')
        
        if p1:
            c = ws.cell(row=r_idx, column=3, value=p1)
            c.font = PRICE_FONT; c.number_format = '#,##0'
            if is_clickable_url(url1):
                c.hyperlink = url1
        elif url1 and is_clickable_url(url1):
            c = ws.cell(row=r_idx, column=3, value='Цена не указана')
            c.font = Font(name='Calibri', size=9, color='9C5700', italic=True)
            c.hyperlink = url1
        else:
            c = ws.cell(row=r_idx, column=3, value='Не найдена')
            c.font = Font(name='Calibri', size=9, color='9C0006', italic=True)
        
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        
        # Shop 1
        s1 = item.get('supplier1', '—')
        c = ws.cell(row=r_idx, column=4, value=s1)
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        c.font = SHOP_FONT if p1 else Font(name='Calibri', size=8, color='999999', italic=True)
        
        # URL 1
        url1_val = item.get('url1', '')
        c = ws.cell(row=r_idx, column=5, value=url1_val if is_clickable_url(url1_val) else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        if is_clickable_url(url1_val):
            c.font = LINK_FONT
            c.hyperlink = url1_val
        else:
            c.font = Font(name='Calibri', size=8, color='999999', italic=True)
        
        # === Price 2 (optional, 10 min limit) ===
        p2 = item.get('price2')
        url2 = item.get('url2', '')
        
        if p2:
            c = ws.cell(row=r_idx, column=6, value=p2)
            c.font = PRICE_FONT; c.number_format = '#,##0'
            if is_clickable_url(url2):
                c.hyperlink = url2
        elif url2 and is_clickable_url(url2):
            c = ws.cell(row=r_idx, column=6, value='Цена не указана')
            c.font = Font(name='Calibri', size=9, color='9C5700', italic=True)
            c.hyperlink = url2
        else:
            c = ws.cell(row=r_idx, column=6, value='Не найдена')
            c.font = Font(name='Calibri', size=9, color='9C5700', italic=True)
        
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        
        # Shop 2
        s2 = item.get('supplier2', '—')
        c = ws.cell(row=r_idx, column=7, value=s2)
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        c.font = SHOP_FONT if p2 else Font(name='Calibri', size=8, color='999999', italic=True)
        
        # URL 2
        url2_val = item.get('url2', '')
        c = ws.cell(row=r_idx, column=8, value=url2_val if is_clickable_url(url2_val) else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        if is_clickable_url(url2_val):
            c.font = LINK_FONT
            c.hyperlink = url2_val
        else:
            c.font = Font(name='Calibri', size=8, color='999999', italic=True)
        
        # === Analog other brand (optional, 5 min limit) ===
        pa = item.get('analog_price')
        analog_url = item.get('analog_url', '')
        
        if pa:
            c = ws.cell(row=r_idx, column=9, value=pa)
            c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
            c.font = PRICE_FONT; c.number_format = '#,##0'
            if is_clickable_url(analog_url):
                c.hyperlink = analog_url
        else:
            c = ws.cell(row=r_idx, column=9, value='—')
            c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
            c.font = Font(name='Calibri', size=10, color='999999')
        
        # Analog shop
        a_s = item.get('analog_supplier', '—')
        c = ws.cell(row=r_idx, column=10, value=a_s if pa else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        c.font = SHOP_FONT if pa else Font(name='Calibri', size=8, color='999999', italic=True)
        
        # Analog URL
        c = ws.cell(row=r_idx, column=11, value=analog_url if is_clickable_url(analog_url) else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        if is_clickable_url(analog_url):
            c.font = LINK_FONT
            c.hyperlink = analog_url
        else:
            c.font = Font(name='Calibri', size=8, color='999999', italic=True)
        
        # === Alternative same brand (optional, 10 min limit) ===
        alt_p = item.get('alt_price')
        alt_url = item.get('alt_url', '')
        
        if alt_p:
            c = ws.cell(row=r_idx, column=12, value=alt_p)
            c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
            c.font = PRICE_FONT; c.number_format = '#,##0'
            if is_clickable_url(alt_url):
                c.hyperlink = alt_url
        else:
            c = ws.cell(row=r_idx, column=12, value='—')
            c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
            c.font = Font(name='Calibri', size=10, color='999999')
        
        # Alt shop
        alt_s = item.get('alt_supplier', '—')
        c = ws.cell(row=r_idx, column=13, value=alt_s if alt_p else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        c.font = SHOP_FONT if alt_p else Font(name='Calibri', size=8, color='999999', italic=True)
        
        # Alt URL
        c = ws.cell(row=r_idx, column=14, value=alt_url if is_clickable_url(alt_url) else '—')
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER; c.fill = bg
        if is_clickable_url(alt_url):
            c.font = LINK_FONT
            c.hyperlink = alt_url
        else:
            c.font = Font(name='Calibri', size=8, color='999999', italic=True)
        
        # === Recommendation ===
        rec = item.get('comment', '')
        c = ws.cell(row=r_idx, column=15, value=rec)
        c.alignment = CENTER_ALIGN; c.border = THIN_BORDER
        c.font = Font(name='Calibri', size=9, bold=True)
        
        if rec:
            if '🔴' in rec or 'НЕ' in rec:
                c.fill = REC_RED
                c.font = Font(name='Calibri', size=9, bold=True, color='9C0006')
            elif '⚠️' in rec or 'Требует' in rec:
                c.fill = REC_YELLOW
                c.font = Font(name='Calibri', size=9, bold=True, color='9C5700')
            else:
                c.fill = REC_GREEN
                c.font = Font(name='Calibri', size=9, bold=True, color='006100')
        elif not p1:
            c.value = '❌ Цена 1 не найдена — позиция не оценена'
            c.fill = REC_RED
            c.font = Font(name='Calibri', size=9, bold=True, color='9C0006')
        elif not p2 and not pa and not alt_p:
            c.value = 'Оригинал подтверждён. Аналоги не найдены.'
            c.fill = REC_GREEN
            c.font = Font(name='Calibri', size=9, bold=True, color='006100')
        elif not p2:
            c.value = 'Цена 1 подтверждена. Второй поставщик не найден.'
            c.fill = REC_YELLOW
            c.font = Font(name='Calibri', size=9, bold=True, color='9C5700')
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 28
    
    ws.freeze_panes = 'A2'
    
    # Prepare analogs for sub-agent research
    analogs_to_research = []
    for item in results:
        if item.get('analog_brand'):
            analogs_to_research.append({
                'type': 'other_brand',
                'original': item['name'],
                'analog': item['analog_brand'],
                'original_price': item.get('price1'),
                'analog_price': item.get('analog_price'),
                'row_num': item.get('num')
            })
        if item.get('alt_brand'):
            analogs_to_research.append({
                'type': 'same_brand',
                'original': item['name'],
                'analog': item['alt_brand'],
                'original_price': item.get('price1'),
                'analog_price': item.get('alt_price'),
                'row_num': item.get('num')
            })
    
    return analogs_to_research


def create_analog_matrices(wb, results):
    """Create analog comparison matrices as separate sheets.
    
    Sheets are created for ALL items. If analog not found — show 'not found' message.
    """
    
    # --- Analogs: Other Brand ---
    ws = wb.create_sheet(title='Аналоги другой марки')
    row = 1
    
    for item in results:
        has_analog = bool(item.get('analog_brand'))
        
        # Section title
        if has_analog:
            title = f"№{item['num']}: {item['name'][:40]} → {item['analog_brand']}"
        else:
            title = f"№{item['num']}: {item['name'][:40]} → Аналоги не найдены"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_cell.alignment = CENTER_ALIGN
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        
        if has_analog:
            # Headers
            headers = ['Параметр', 'Оригинал', 'Аналог', 'Отклонение', 'Влияние']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            row += 1
            
            # Price row: handle null analog_price
            if item.get('analog_price'):
                price_analog = f"{item['analog_price']} ₽"
                if item.get('price1') and item['price1'] > 0:
                    savings = int((1 - item['analog_price']/item['price1'])*100)
                    influence = f"Экономия {savings}%"
                else:
                    influence = "Экономия: н/д (нет цены оригинала)"
            else:
                price_analog = "По запросу"
                influence = "Цену уточняйте у поставщика"
            
            # Data rows — dynamic based on available specs
            # Build rows from actual data, fallback to minimal set
            data_rows = [
                ['Цена', f"{item.get('price1', '—')} ₽", price_analog, '—', influence],
                ['Производитель', item['name'].split()[0] if item['name'] else '—', 
                 item['analog_brand'].split()[0] if item.get('analog_brand') else '—', 
                 'Другая марка' if item.get('analog_brand') else '—', 
                 'Сравнить спецификации'],
            ]
            
            # Add spec rows if available in results
            specs = item.get('specs', {})
            if specs and isinstance(specs, dict):
                for param, value in specs.items():
                    analog_val = specs.get(f"analog_{param}", value)
                    match = '✅ Совпадает' if str(value).lower() == str(analog_val).lower() else '⚠️ Проверить'
                    data_rows.append([param, value, analog_val, match, '—'])
            
            # Fallback minimal spec rows if no specs provided
            if len(data_rows) < 3:
                data_rows.extend([
                    ['Описание', item.get('description', 'По спецификации') or 'По спецификации', 
                     'По спецификации', '—', 'Сравнить характеристики'],
                ])
            
            for row_data in data_rows:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGN
                    cell.font = Font(name='Calibri', size=9)
                row += 1
            
            # Sources
            if item.get('analog_url'):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                src_cell = ws.cell(row=row, column=1, value=f"Источник: {item['analog_url']}")
                src_cell.fill = GRAY_FILL
                src_cell.font = Font(name='Calibri', size=8, color='0563C1', underline='single')
                src_cell.hyperlink = item['analog_url']
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
            else:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                src_cell = ws.cell(row=row, column=1, value="Источник: не указан — уточните цену у поставщика (Human-in-the-loop)")
                src_cell.fill = GRAY_FILL
                src_cell.font = Font(name='Calibri', size=8, italic=True, color='996600')
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
        else:
            # No analog found
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            msg_cell = ws.cell(row=row, column=1, value="Аналоги другой марки не найдены")
            msg_cell.fill = GRAY_FILL
            msg_cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
            msg_cell.alignment = CENTER_ALIGN
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        
        row += 1  # Empty row between sections
    
    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    
    # --- Analogs: Same Brand ---
    ws = wb.create_sheet(title='Аналоги той же марки')
    row = 1
    
    for item in results:
        has_alt = bool(item.get('alt_brand'))
        
        # Section title
        if has_alt:
            title = f"№{item['num']}: {item['name'][:40]} → {item['alt_brand']}"
        else:
            title = f"№{item['num']}: {item['name'][:40]} → Альтернативы не найдены"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        title_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_cell.alignment = CENTER_ALIGN
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
        
        if has_alt:
            # Headers
            headers = ['Параметр', 'Оригинал', 'Альтернатива', 'Отклонение', 'Влияние']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            row += 1
            
            # Price row: handle null alt_price
            if item.get('alt_price'):
                price_alt = f"{item['alt_price']} ₽"
                if item.get('price1') and item['price1'] > 0:
                    diff = int((item['alt_price']/item['price1']-1)*100)
                    influence = f"Разница {diff}%"
                else:
                    influence = "Разница: н/д (нет цены оригинала)"
            else:
                price_alt = "По запросу"
                influence = "Цену уточняйте у поставщика"
            
            # Data rows — dynamic based on available specs
            data_rows = [
                ['Цена', f"{item.get('price1', '—')} ₽", price_alt, '—', influence],
                ['Производитель / Модель', 
                 item['name'].split()[0] if item['name'] else '—',
                 item['alt_brand'].split()[0] if item.get('alt_brand') else '—',
                 'Другая модель' if item.get('alt_brand') else '—',
                 'Та же марка, другая модель'],
            ]
            
            # Add spec rows if available
            alt_specs = item.get('alt_specs', {})
            if alt_specs and isinstance(alt_specs, dict):
                for param, value in alt_specs.items():
                    original_val = item.get('specs', {}).get(param, '—') if item.get('specs') else '—'
                    match = '✅ Совпадает' if str(original_val).lower() == str(value).lower() else '⚠️ Отличается'
                    data_rows.append([param, original_val, value, match, '—'])
            
            # Fallback minimal spec rows
            if len(data_rows) < 3:
                data_rows.extend([
                    ['Описание', item.get('description', 'По спецификации') or 'По спецификации',
                     'По спецификации', '—', 'Сравнить характеристики'],
                ])
            
            for row_data in data_rows:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGN
                    cell.font = Font(name='Calibri', size=9)
                row += 1
            
            # Sources
            if item.get('alt_url'):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                src_cell = ws.cell(row=row, column=1, value=f"Источник: {item['alt_url']}")
                src_cell.fill = GRAY_FILL
                src_cell.font = Font(name='Calibri', size=8, color='0563C1', underline='single')
                src_cell.hyperlink = item['alt_url']
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
            else:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                src_cell = ws.cell(row=row, column=1, value="Источник: не указан — уточните цену у поставщика (Human-in-the-loop)")
                src_cell.fill = GRAY_FILL
                src_cell.font = Font(name='Calibri', size=8, italic=True, color='996600')
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
        else:
            # No alt found
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            msg_cell = ws.cell(row=row, column=1, value="Аналоги внутри марки не найдены")
            msg_cell.fill = GRAY_FILL
            msg_cell.font = Font(name='Calibri', size=10, italic=True, color='666666')
            msg_cell.alignment = CENTER_ALIGN
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
        
        row += 1  # Empty row between sections
    
    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 runner_v3.py input.xlsx results.json [output_dir]")
        print("\nresults.json format (2+1+1):")
        print(json.dumps([{
            "num": 1, "name": "Product Name",
            "price1": 100000, "supplier1": "supplier.ru", "url1": "https://...",
            "price2": null, "supplier2": "", "url2": "",
            "analog_brand": "Analog Brand", "analog_price": 80000,
            "analog_supplier": "...", "analog_url": "",
            "alt_brand": "Alt Model", "alt_price": null,
            "alt_supplier": "", "alt_url": "",
            "comment": "Согласовать...", "date": "2026-09-01"
        }], indent=2, ensure_ascii=False))
        sys.exit(1)
    
    input_file = sys.argv[1]
    results_file = sys.argv[2]
    output_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.dirname(input_file) or '.'
    
    with open(results_file, 'r', encoding='utf-8') as f:
        results = json.load(f)
    
    print(f"Loaded {len(results)} items from {results_file}")
    
    # Inline eval
    print("\n🔍 Running inline eval...")
    eval_result = inline_eval_all(results)
    if not eval_result['passed']:
        print("\n❌ Inline eval FAILED. Fix errors before proceeding.")
        print(f"   Total errors: {eval_result['total_errors']}")
        sys.exit(1)
    print("✅ Inline eval passed\n")
    
    wb = openpyxl.Workbook()
    analogs_to_research = create_main_sheet(wb, results)
    create_analog_matrices(wb, results)  # Add analog matrix sheets
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    intermediate_file = os.path.join(output_dir, f'price_comparison_main_{timestamp}.xlsx')
    wb.save(intermediate_file)
    
    # Separate files for each analog type
    analogs_other = [a for a in analogs_to_research if a['type'] == 'other_brand']
    analogs_same = [a for a in analogs_to_research if a['type'] == 'same_brand']
    
    analogs_file_other = os.path.join(output_dir, f'analogs_other_brand_{timestamp}.json')
    analogs_file_same = os.path.join(output_dir, f'analogs_same_brand_{timestamp}.json')
    
    with open(analogs_file_other, 'w', encoding='utf-8') as f:
        json.dump(analogs_other, f, ensure_ascii=False, indent=2)
    with open(analogs_file_same, 'w', encoding='utf-8') as f:
        json.dump(analogs_same, f, ensure_ascii=False, indent=2)
    
    print(f"Main results saved: {intermediate_file}")
    print(f"Analogs (other brand): {analogs_file_other} ({len(analogs_other)} items)")
    print(f"Analogs (same brand): {analogs_file_same} ({len(analogs_same)} items)")
    print(f"Total analogs to research: {len(analogs_to_research)}")


if __name__ == '__main__':
    main()
