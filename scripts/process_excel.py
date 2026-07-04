#!/usr/bin/env python3
"""
Excel processor for price comparison skill.
Adds columns for supplier prices and contacts.

Usage:
    python3 process_excel.py input.xlsx [output.xlsx]

If output is not specified, creates input_с_ценами.xlsx
"""

import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install: pip install openpyxl")
    sys.exit(1)


def process_excel(input_path, output_path=None):
    """Add price comparison columns to Excel file."""
    
    input_path = Path(input_path)
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)
    
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_с_ценами{input_path.suffix}"
    else:
        output_path = Path(output_path)
    
    # Load workbook
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Find the header row (assume first row with data)
    header_row = 1
    max_col = ws.max_column
    
    # Add new columns after existing ones
    new_cols = [
        "Цена поставщик 1",
        "Сайт поставщика 1",
        "Цена поставщик 2", 
        "Сайт поставщика 2",
        "Email поставщика (если нет цен)"
    ]
    
    for i, col_name in enumerate(new_cols):
        col_num = max_col + i + 1
        cell = ws.cell(row=header_row, column=col_num, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths for new columns
    for i in range(len(new_cols)):
        col_letter = openpyxl.utils.get_column_letter(max_col + i + 1)
        ws.column_dimensions[col_letter].width = 25
    
    # Save
    wb.save(output_path)
    print(f"Saved to: {output_path}")
    print(f"Added {len(new_cols)} columns for price comparison")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 process_excel.py input.xlsx [output.xlsx]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    process_excel(input_file, output_file)
